from openpyxl import load_workbook
from openpyxl.styles import PatternFill



def ajuster_largeur_colonnes_et_styles(fichier):
    wb = load_workbook(fichier)
    ws = wb.active
    
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for cell in ws['A']:
        #if cell.row != 1:  # Skip the header
        cell.fill = fill_blue

    for col in ws.columns:
        
        max_length = 0
        column = col[0].column_letter
        min_val = float('inf')
        max_val = float('-inf')
        min_cell = None
        max_cell = None
        for cell in col:
            try:
                cell_value = str(cell.value)
                if len(str(cell.value)) > max_length:
                    max_length = len(cell_value)
                if cell.row != 1:  # Skip the header
                    val = float(cell.value.replace('%', ''))
                    if val < min_val:
                        min_val = val
                        min_cell = cell
                    if val > max_val:
                        max_val = val
                        max_cell = cell
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
        if min_cell:
            min_cell.fill = fill_red
        if max_cell:
            max_cell.fill = fill_green

    wb.save(fichier)

